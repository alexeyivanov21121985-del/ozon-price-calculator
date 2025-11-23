import streamlit as st
import pandas as pd
import numpy as np
import io
import tempfile
import os
import openpyxl
import math
import re

try:
    from cost_updater import process_cost_update_logic, to_float
except ImportError:
    st.error("ОШИБКА: Файл 'cost_updater.py' не найден!")
    st.stop()

# ===================== УТИЛИТЫ =====================
def ceil_to_step(x: float, step: float) -> float:
    if step and step > 0:
        return math.ceil(float(x) / step) * step
    return float(x)

def to_fraction(x) -> float:
    f = to_float(x, default=0.0)
    if np.isnan(f) or f < 0: return 0.0
    if f > 1.0: f = f / 100.0
    return max(0.0, min(f, 0.9999))

def get_file_columns(file_obj, header_row_idx, sheet_name=0):
    """Возвращает список колонок"""
    file_obj.seek(0)
    try:
        df = pd.read_excel(file_obj, sheet_name=sheet_name, header=header_row_idx, nrows=0)
        return list(df.columns)
    except Exception:
        return []

def detect_header(file_obj, keywords, sheet_name=0):
    """Автопоиск строки заголовка"""
    file_obj.seek(0)
    try:
        df = pd.read_excel(file_obj, sheet_name=sheet_name, header=None, nrows=20)
        for idx, row in df.iterrows():
            s = [str(x).lower() for x in row.values]
            if sum(1 for k in keywords if any(k in val for val in s)) >= 1:
                return idx
    except: pass
    return 0

# ===================== ЛОГИКА КАЛЬКУЛЯТОРА (Вкладка 2) =====================
def process_selling_price(uploaded_file, config):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.getvalue())
        tmp_path = tmp.name

    try:
        # Читаем с указанным хедером
        df = pd.read_excel(tmp_path, sheet_name=config['sheet_name'], header=config['header_row_idx'], engine='openpyxl')
        
        # Функция поиска индекса по имени (для калькулятора пока оставляем по имени, т.к. это второй этап)
        def get_col_idx(name):
            if name in df.columns: return df.columns.get_loc(name)
            return None # Упрощено, так как пользователь выбирает из списка
        
        idx_cost = get_col_idx(config['col_cost'])
        idx_ozon = get_col_idx(config['col_ozon'])
        
        idx_fixes = []
        for f in config['col_fix']:
            ix = get_col_idx(f)
            if ix is not None: idx_fixes.append(ix)

        results = []
        for i, row in df.iterrows():
            # Используем iloc для надежности
            try:
                cost = to_float(row.iloc[idx_cost])
                ozon = to_fraction(row.iloc[idx_ozon])
                fix = sum(to_float(row.iloc[fix_ix]) for fix_ix in idx_fixes)
                
                def calc(margin):
                    denom = 1 - (ozon + margin)
                    if denom <= 0.01 or np.isnan(cost): return None
                    res = (cost + fix) / denom
                    return ceil_to_step(res, config['round_step'])

                p_full = calc(config['margin_std'])
                p_min = calc(config['margin_min'])
                
                if p_full or p_min:
                    results.append({'idx': i, 'full': p_full, 'min': p_min})
            except: continue
                
        idx_dest_full = get_col_idx(config['dest_full'])
        idx_dest_min = get_col_idx(config['dest_min'])

        wb = openpyxl.load_workbook(tmp_path)
        ws = wb[config['sheet_name']]
        start_row = config['header_row_idx'] + 2
        
        cnt = 0
        for item in results:
            r = start_row + item['idx']
            if item['full'] and idx_dest_full is not None: 
                ws.cell(row=r, column=idx_dest_full+1).value = item['full']
            if item['min'] and idx_dest_min is not None: 
                ws.cell(row=r, column=idx_dest_min+1).value = item['min']
            cnt += 1
            
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        wb.close()
        os.unlink(tmp_path)
        return out, cnt
    except Exception as e:
        if os.path.exists(tmp_path): os.unlink(tmp_path)
        raise e

# ===================== ИНТЕРФЕЙС =====================
st.set_page_config(page_title="Ozon Master 3.0", layout="wide")
st.title("Ozon Master Tool 3.0 (Fix)")

tab1, tab2 = st.tabs(["1. СЕБЕСТОИМОСТЬ", "2. ЦЕНЫ ПРОДАЖИ"])

# --- TAB 1 ---
with tab1:
    st.info("Шаг 1: Обновление себестоимости. Работает по индексам колонок.")
    
    col_f1, col_f2 = st.columns(2)
    ozon_file = col_f1.file_uploader("Файл Ozon", type=["xlsx"], key="o1")
    price_file = col_f2.file_uploader("Прайс-лист", type=["xlsx"], key="p1")

    if ozon_file and price_file:
        st.divider()
        
        # 1. Лист Ozon
        wb_temp = openpyxl.load_workbook(ozon_file, read_only=True, data_only=True)
        ozon_sheets = wb_temp.sheetnames
        wb_temp.close()
        
        def_sh = 0
        for i, s in enumerate(ozon_sheets):
            if "товар" in s.lower() or "цены" in s.lower(): def_sh = i; break
            
        c_sh1, c_sh2 = st.columns(2)
        with c_sh1:
            sel_ozon_sheet = st.selectbox("Лист Ozon", ozon_sheets, index=def_sh)
            
        # 2. Строки заголовков
        c_r1, c_r2 = st.columns(2)
        with c_r1:
            # Автопоиск
            h_ozon_user = st.number_input("Строка заголовка Ozon (обычно 2)", min_value=1, value=2, step=1)
            h_ozon_idx = h_ozon_user - 1
        with c_r2:
            auto_p = detect_header(price_file, ["код", "цена", "артикул"])
            h_price_user = st.number_input("Строка заголовка Прайса", min_value=1, value=auto_p+1, step=1)
            h_price_idx = h_price_user - 1
            
        # 3. Чтение колонок
        cols_ozon = get_file_columns(ozon_file, h_ozon_idx, sheet_name=sel_ozon_sheet)
        cols_price = get_file_columns(price_file, h_price_idx)
        
        st.divider()
        
        if cols_ozon and cols_price:
            c_sel1, c_sel2 = st.columns(2)
            
            # Helper для поиска индекса
            def find_i(lst, keys):
                for i, c in enumerate(lst):
                    if any(k in str(c).lower() for k in keys): return i
                return 0

            with c_sel1:
                st.markdown("**Настройки Ozon**")
                val_art_o = st.selectbox("Колонка 'Артикул'", cols_ozon, index=find_i(cols_ozon, ["артикул"]))
                val_cost_o = st.selectbox("Куда писать 'Себестоимость'", cols_ozon, index=find_i(cols_ozon, ["себест", "закуп"]))
                
                # Получаем индексы выбранных элементов
                idx_art_o_final = cols_ozon.index(val_art_o)
                idx_cost_o_final = cols_ozon.index(val_cost_o)
                
            with c_sel2:
                st.markdown("**Настройки Прайса**")
                val_code_p = st.selectbox("Код товара", cols_price, index=find_i(cols_price, ["код"]))
                val_price_p = st.selectbox("Цена за шт.", cols_price, index=find_i(cols_price, ["цена"]))
                val_qty_p = st.selectbox("Кол-во в упаковке", cols_price, index=find_i(cols_price, ["упак", "кол", "штук"]))
                
                idx_code_p_final = cols_price.index(val_code_p)
                idx_price_p_final = cols_price.index(val_price_p)
                idx_qty_p_final = cols_price.index(val_qty_p)
            
            st.divider()
            if st.button("🚀 Проверить и Обновить", type="primary"):
                cfg = {
                    'ozon_sheet_name': sel_ozon_sheet,
                    'header_ozon_idx': h_ozon_idx,
                    'header_price_idx': h_price_idx,
                    'col_idx_art_ozon': idx_art_o_final,
                    'col_idx_cost_ozon': idx_cost_o_final,
                    'col_idx_code_price': idx_code_p_final,
                    'col_idx_price_price': idx_price_p_final,
                    'col_idx_qty_price': idx_qty_p_final
                }
                
                with st.spinner("Анализируем..."):
                    try:
                        ozon_file.seek(0)
                        price_file.seek(0)
                        # Теперь функция возвращает еще и DataFrame для отладки
                        res, miss, fnd, msd, debug_df = process_cost_update_logic(ozon_file, price_file, cfg)
                        
                        st.subheader("🔍 Проверка расчетов (Первые 500 строк)")
                        st.caption("Убедитесь, что 'Цена (Прайс)' и 'Кол-во (Итог)' определились верно.")
                        st.dataframe(debug_df, use_container_width=True)
                        
                        st.success(f"Обработано: {fnd} строк. Пропущено: {msd}.")
                        
                        b1, b2 = st.columns(2)
                        b1.download_button("📥 Скачать Ozon с себестоимостью", res, f"Cost_{ozon_file.name}", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        if miss: b2.download_button("📥 Скачать список ненайденных", miss, "missing.xlsx")
                        
                    except Exception as e:
                        st.error(f"Ошибка: {e}")
        else:
            st.error("Колонки не найдены. Попробуйте изменить номера строк заголовков выше.")

# --- TAB 2 ---
with tab2:
    st.info("Шаг 2: Расчет цен (Выбираем файл с Шага 1)")
    u2 = st.file_uploader("Файл с себестоимостью", type=["xlsx"], key="u2")
    if u2:
        # ... (Стандартная логика UI для таба 2)
        wb = openpyxl.load_workbook(u2, read_only=True, data_only=True)
        shs = wb.sheetnames
        wb.close()
        def_s = 0
        for i, s in enumerate(shs):
            if "товар" in s.lower(): def_s = i; break
        sh2 = st.selectbox("Лист", shs, index=def_s, key="sh2")
        
        ha = detect_header(u2, ["артикул", "цена"], sheet_name=sh2)
        h2 = st.number_input("Строка заголовка", min_value=1, value=ha+1, step=1, key="h2_in")
        h_ix2 = h2 - 1
        
        cols2 = get_file_columns(u2, h_ix2, sheet_name=sh2)
        
        if cols2:
            c1, c2 = st.columns(2)
            def fi(k):
                for i,c in enumerate(cols2):
                    if k in str(c).lower(): return i
                return 0
            with c1:
                ms = st.number_input("Маржа %", 20.0)
                mm = st.number_input("Мин %", 10.0)
                rs = st.number_input("Шаг", 10.0)
            with c2:
                vc = st.selectbox("Себест.", cols2, index=fi("себест"))
                vo = st.selectbox("Ozon", cols2, index=fi("ozon"))
                vf = st.multiselect("Фиксы", cols2, default=[c for c in cols2 if "лог" in str(c).lower()])
                vdf = st.selectbox("Запись: Цена", cols2, index=fi("новая цена"))
                vdm = st.selectbox("Запись: Мин", cols2, index=fi("минимальная"))
                
            if st.button("Рассчитать", type="primary"):
                cfg2 = {
                    'sheet_name': sh2, 'header_row_idx': h_ix2,
                    'margin_std': ms/100, 'margin_min': mm/100, 'round_step': rs,
                    'col_cost': vc, 'col_ozon': vo, 'col_fix': vf,
                    'dest_full': vdf, 'dest_min': vdm
                }
                try:
                    u2.seek(0)
                    rf, cnt = process_selling_price(u2, cfg2)
                    st.success(f"Готово: {cnt}")
                    st.download_button("Скачать", rf, f"Final_{u2.name}")
                except Exception as e: st.error(f"Err: {e}")