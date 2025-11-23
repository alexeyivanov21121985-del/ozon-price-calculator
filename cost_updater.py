import pandas as pd
import numpy as np
import re
import io
import os
import tempfile
import openpyxl

def to_float(v, default=np.nan) -> float:
    """Превращает '1 200,00' или '1,200.00' в число 1200.0"""
    if v is None: return default
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    # Оставляем только цифры, точку
    s = re.sub(r"[^\d\.]", "", s)
    try:
        return float(s) if s else default
    except:
        return default

def parse_article_quantity(ozon_article):
    """
    Разбирает 'УТ-001(10)' -> код 'УТ-001', кол-во 10.
    Если скобок нет -> код 'УТ-001', кол-во None.
    """
    s_art = str(ozon_article).strip()
    match = re.search(r'^(.*?)\((\d+)\)$', s_art)
    if match:
        clean_code = match.group(1).strip()
        qty = int(match.group(2))
        return clean_code, qty
    return s_art, None

def process_cost_update_logic(ozon_file, price_file, config):
    # ================= 1. ЧИТАЕМ ПРАЙС =================
    # Читаем как строки, чтобы коды не ломались
    df_price = pd.read_excel(price_file, header=config['header_price_idx'], dtype=str)
    
    # Получаем индексы (номера) колонок, которые выбрал пользователь
    idx_code = config['col_idx_code_price']
    idx_price = config['col_idx_price_price']
    idx_qty = config['col_idx_qty_price']

    # Создаем справочник цен
    # Ключ: Код товара, Значение: {Цена, Кол-во_в_упаковке}
    price_map = {}
    for _, row in df_price.iterrows():
        # Берем данные строго по номеру колонки (iloc)
        try:
            code_val = str(row.iloc[idx_code]).strip()
            price_raw = row.iloc[idx_price]
            qty_raw = row.iloc[idx_qty]
            
            price_val = to_float(price_raw)
            qty_val = to_float(qty_raw)
            
            # Если кол-во не указано или 0, считаем как 1
            if np.isnan(qty_val) or qty_val <= 0:
                qty_val = 1.0
            
            if code_val and not np.isnan(price_val):
                price_map[code_val] = {'price': price_val, 'pack_qty': qty_val}
        except IndexError:
            continue

    # ================= 2. ГОТОВИМ ФАЙЛ OZON =================
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(ozon_file.getvalue())
        tmp_path = tmp.name

    try:
        # Читаем Ozon
        sh_name = config.get('ozon_sheet_name', 0)
        df_ozon = pd.read_excel(tmp_path, sheet_name=sh_name, header=config['header_ozon_idx'], dtype=str)
        
        idx_art_o = config['col_idx_art_ozon']
        # Индекс колонки для записи в Excel (1-based) = индекс в pandas (0-based) + 1
        target_col_excel = config['col_idx_cost_ozon'] + 1
        
        # Для отладки собираем таблицу
        debug_data = []
        
        # Открываем Excel на запись
        wb = openpyxl.load_workbook(tmp_path)
        if isinstance(sh_name, str) and sh_name in wb.sheetnames:
            ws = wb[sh_name]
        else:
            ws = wb.active

        # Вычисляем строку начала данных: Заголовок(0-based) + 1(Excel) + 1(След.строка)
        start_data_row = config['header_ozon_idx'] + 2
        
        found_count = 0
        missing_rows = []

        for i, row in df_ozon.iterrows():
            try:
                raw_art = row.iloc[idx_art_o]
            except IndexError:
                continue
                
            if pd.isna(raw_art): continue
            
            # 1. Парсим артикул
            clean_code, bracket_qty = parse_article_quantity(raw_art)
            
            # 2. Ищем в прайсе
            prod_data = price_map.get(clean_code)
            
            status = "Не найден"
            final_cost = 0
            used_price = 0
            used_qty = 0
            calc_formula = ""
            
            if prod_data:
                status = "OK"
                price = prod_data['price']
                pack_qty = prod_data['pack_qty']
                
                # ЛОГИКА ВЫБОРА КОЛИЧЕСТВА
                if bracket_qty is not None:
                    # Если есть скобки (10), берем их
                    final_qty = bracket_qty
                    calc_formula = f"{price} * {final_qty} (из скобок)"
                else:
                    # Если скобок нет, берем из колонки прайса
                    final_qty = pack_qty
                    calc_formula = f"{price} * {final_qty} (из прайса)"
                
                final_cost = price * final_qty
                
                # Пишем в Excel
                r_idx = start_data_row + i
                ws.cell(row=r_idx, column=target_col_excel).value = final_cost
                found_count += 1
                
                used_price = price
                used_qty = final_qty
            else:
                missing_rows.append({'Артикул': raw_art, 'Код поиска': clean_code})

            # Добавляем в отчет (первые 100 строк или все, чтобы не перегружать память)
            if len(debug_data) < 500: 
                debug_data.append({
                    "Артикул Ozon": raw_art,
                    "Код поиска": clean_code,
                    "Статус": status,
                    "Цена (Прайс)": used_price,
                    "Кол-во (Итог)": used_qty,
                    "Расчет": calc_formula,
                    "Итог Себест.": final_cost
                })

        out_ozon = io.BytesIO()
        wb.save(out_ozon)
        out_ozon.seek(0)
        wb.close()
        os.unlink(tmp_path)
        
        out_missing = None
        if missing_rows:
            df_miss = pd.DataFrame(missing_rows)
            out_missing = io.BytesIO()
            with pd.ExcelWriter(out_missing, engine='openpyxl') as writer:
                df_miss.to_excel(writer, index=False)
            out_missing.seek(0)
            
        return out_ozon, out_missing, found_count, len(missing_rows), pd.DataFrame(debug_data)

    except Exception as e:
        if os.path.exists(tmp_path): os.unlink(tmp_path)
        raise e